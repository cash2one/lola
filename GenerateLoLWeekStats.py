
# coding: utf-8

# In[1]:

import csv
import codecs
import requests
import json
import urllib
import os
from openpyxl import load_workbook
from bs4 import BeautifulSoup


# In[98]:

RESULT_FOLDER = "./Results/"

lpl_match_file_name = "./2017LPL_Match.csv"
lpl_player_file_name = "./2017LPL_Player.csv"
output_player_file_name = "Week%sStats_Players.csv"
output_champion_file_name = "Week%sStats_Champions.csv"
week_stats_file_name = "Week%sStats.csv"


# In[ ]:

print "从oracleselixir.com下载数据"
file_opener = urllib.URLopener()
(downloaed_filename, headers) = file_opener.retrieve("http://oracleselixir.com/gamedata/2017-spring/")


# In[4]:

for a_item in headers.headers:
    if a_item.startswith("Content-Disposition: attachment; filename"):
        filename = a_item.split("\"")[1]
        break

filedate = filename.split('.')[0].split('-')[-3:]

print "当前数据日期是%s年%s月%s日" % (filedate[0] ,filedate[1], filedate[2])

if not os.path.exists(RESULT_FOLDER):
    os.makedirs(RESULT_FOLDER)

os.rename(downloaed_filename, RESULT_FOLDER + filename)
print "从oracleselixir.com下载的数据已保存至%s" % RESULT_FOLDER + filename


# In[20]:

workbook = load_workbook(RESULT_FOLDER + filename, use_iterators=True)
first_sheet = workbook.get_sheet_names()[0]
worksheet = workbook.get_sheet_by_name(first_sheet)

REGION_COLUMN = 2
WEEKDAY_COLUMN = 5

#检查LCK赛区在第几周，由于各赛区周数不同，这里以LCK赛区为准
latest_week = 0
row_number = worksheet.max_row
for row_idx, row in enumerate(worksheet.iter_rows()):
    league = row[REGION_COLUMN].value
    week_day = str(row[WEEKDAY_COLUMN].value)
    week = week_day.split('.')[0]
    if league == "LCK":
        latest_week = week
    if row_idx % 1000 == 0:
        print "正在检查比赛数据%d/%d" % (row_idx, row_number)
print "LCK赛区最新数据为第%s周" % latest_week

# check out the last row
#for cell in row:
#    print cell


# In[80]:

#将最新周数据保存至独立的csv文件
filename_splits = filename.split('.')
week_file = (filename_splits[0] + "_week%s" + ".csv") % (latest_week)

print "将第%s周的数据写入csv文件%s" % (latest_week, RESULT_FOLDER + week_file)
rewrite_start = False
with open(RESULT_FOLDER + week_file, 'wb') as fd:
    for row_idx, row in enumerate(worksheet.iter_rows()):
        if rewrite_start == False and row_idx != 0:
            league = row[REGION_COLUMN].value
            week_day = str(row[WEEKDAY_COLUMN].value)
            week = week_day.split('.')[0]
            if row_idx % 1000 == 0:
                print "略过比赛数据%d/%d" % (row_idx, row_number)
            if league == "LCK" and week == latest_week:
                rewrite_start = True
            else:
                continue
        values=[] 
        for cell in row: 
            value=cell.value 
            if value is None: 
                value='' 
            values.append(str(value))
            #if not isinstance(value, unicode):
            #    value=unicode(value)
            #    value=value.encode('utf8')  
            
        if row_idx % 1000 == 0:
            print "正在写入比赛数据%d/%d" % (row_idx, row_number)
        fd.write(','.join(values))
        fd.write('\n')
print "数据写入完成"


# In[81]:

#读取LPL选手数据
print "读取LPL选手数据"
f = codecs.open(lpl_player_file_name, 'r', 'gb2312', errors='ignore')
lines = f.readlines()
#print "All lines number:" + str(len(lines))
lpl_table = []
lpl_define_row = []
for a_idx, a_line in enumerate(lines):
    words = a_line.split(',')
    if a_idx == 0:
        lpl_define_row = words
    if words[0].startswith('2017'):
        lpl_table.append(words)
#print "Data lines number:" + str(len(lpl_table))


# In[95]:

#读取LPL数据的辅助函数
def TranslateNumber(number):
    chinese_numbers = {1:"一", 2:"二", 3:"三", 4:"四", 5:"五", 6:"六", 7:"七", 8:"八", 9:"九"}
    return chinese_numbers[number]

def TranslateWeek(week_number):
    return "第" + TranslateNumber(week_number) + "周"

def TranslateGame(game_number):
    return "第" + TranslateNumber(game_number) + "场"

chinese_positions = {"ADC":"ADC", "Jungle":"JUNGLE", "Middle":"MID", "Top":"TOP", "Support":"SUPPORT"}
chinese_teams = {"Edward Gaming":"EDG", "Team WE":"WE",                  "Royal Never Give Up":"RNG", "I May":"IM",                  "Invictus Gaming":"IG", "OMG":"OMG",                  "Newbee":"Newbee", "Vici Gaming":"VG",                  "LGD Gaming":"LGD", "Snake Esports":"Snake",                  "QG Reapers":"QG", "Game Talents":"GT"}

LPL_DAMAGE_TAKEN_COLUMN = 28
LPL_DAMAGE_TO_CHAMPION_COLUMN = 25
LPL_TOTAL_GOLD_COLUMN = 17
LPL_WEEK_COLUMN = 2
LPL_GAME_COLUMN = 3
LPL_TEAM_COLUMN = 6
LPL_POSITION_COLUMN = 8
LPL_VS_COLUMN = 4

def GetLPLDamageTaken(week_number, game_number, team_name, opposing_team_name, position):
    #print "Get Damage Taken for "
    #print week_number, game_number, team_name, opposing_team_name, position
    week_chinese = TranslateWeek(week_number)
    game_chinese = TranslateGame(game_number)
    position_chinese = chinese_positions[position]
    
    for a_row in lpl_table:
        current_week = a_row[LPL_WEEK_COLUMN].encode('utf8')
        if current_week.startswith(week_chinese):
            current_game = a_row[LPL_GAME_COLUMN].encode('utf8')
            if current_game.startswith(game_chinese):
                target_team = chinese_teams[team_name.strip()]
                current_team = a_row[LPL_TEAM_COLUMN].encode('utf8').strip()
                if current_team == target_team:
                    current_position = a_row[LPL_POSITION_COLUMN].encode('utf8')
                    if True:#不再确认位置信息，因为腾讯数据的位置信息经常出错current_position.startswith(position_chinese):
                        current_vs = a_row[LPL_VS_COLUMN].encode('utf8').strip()
                        opposing_team = chinese_teams[opposing_team_name.strip()]
                        #print current_vs
                        #print current_team + " vs " + opposing_team
                        if current_vs == current_team + " vs " + opposing_team or current_vs == opposing_team + " vs " + current_team:
                            return (float(a_row[LPL_DAMAGE_TO_CHAMPION_COLUMN]), 
                                    float(a_row[LPL_DAMAGE_TAKEN_COLUMN]),
                                    float(a_row[LPL_TOTAL_GOLD_COLUMN]))
    return (0, 0, 0)


# In[83]:

all_region_define_row = None
all_region_table = []

PLAYERID_COLUMN = 8
print "开始统计全区数据"
with open(RESULT_FOLDER + week_file, 'rb') as csvfile:
    result = csv.reader(csvfile, delimiter=',', quotechar='|')
    for row_idx, row in enumerate(result):
        if row_idx == 0:
            all_region_define_row = row
        else:
            if row[PLAYERID_COLUMN] != "100" and row[PLAYERID_COLUMN] != "200":
                all_region_table.append(row)


# In[96]:

PLAYERID_COLUMN = 8
DPM_COLUMN = 59 + 2 #每分钟对英雄伤害
GPM_COLUMN = 74 #每分钟金钱
DAMAGE_TO_CHAMPIONS_COLUMN = 58 + 2 #对英雄总伤害
KILL_COLUMN = 19 + 2
DEATH_COLUMN = 20 + 2
ASSIST_COLUMN = 21 + 2
LEAGUE_COLUMN = 2
SPLIT_COLUMN = 3
WEEK_COLUMN = 5 #LPL和其余赛区的格式不同：LPL直接是周数，其他赛区是week.day
GAME_COLUMN = 6
URL_COLUMN = 1 #LPL赛区没有这栏信息
SIDE_COLUMN = 9
POSITION_COLUMN = 10
PLAYER_COLUMN = 11
TEAM_COLUMN = 12
CHAMPION_COLUMN = 13
BAN1_COLUMN = 14
BAN2_COLUMN = 15
BAN3_COLUMN = 16
BAN4_COLUMN = 17
BAN5_COLUMN = 18

GAMELENGTH_COLUMN = 17 + 2
RESULT_COLUMN = 18 + 2#0 for lose, 1 for win
TOTAL_GOLD_COLUMN = 73

CUSTOM_WEEK_COLUMN = 95
CUSTOM_DAMAGE_TAKEN_COLUMN = 96

print ("将全区数据统计结果保存至文件" + RESULT_FOLDER + week_stats_file_name) % latest_week
#rewrite these columns to a new file
with open((RESULT_FOLDER + week_stats_file_name) % latest_week, 'wb') as csvfile:
    csvwriter = csv.writer(csvfile, delimiter=',',
                            quotechar='|', quoting=csv.QUOTE_MINIMAL)
    define_row = all_region_define_row
    define_row.append('custom_week')
    define_row.append('custom_damagetaken')
    csvwriter.writerow(define_row)
    total_num = len(all_region_table)
    for a_idx, a_row in enumerate(all_region_table):
        #fill in a neat week number
        old_week = a_row[WEEK_COLUMN]
        current_league = a_row[LEAGUE_COLUMN]
        new_week = old_week
        if current_league != "LPL":
            week_parts = old_week.split('.')
            new_week = week_parts[0]
        a_row.append(new_week)
        
        #get damage taken from tencent data or from match history url
        damage_taken = 0
        current_game = a_row[GAME_COLUMN]
        current_team = a_row[TEAM_COLUMN]
        current_position = a_row[POSITION_COLUMN]
        opposing_team = None
        if a_idx % 10 < 5:
            opposing_team = all_region_table[a_idx + 5][TEAM_COLUMN]
        else:
            opposing_team = all_region_table[a_idx - 5][TEAM_COLUMN]
        if current_league == "LPL":
            (lpl_damage_to_champion, damage_taken, gold_earned) = GetLPLDamageTaken(int(new_week), int(current_game), current_team, opposing_team, current_position)
            #print str(a_idx) + " - LPL damage taken:" + str(damage_taken) + "-" + str(gold_earned)
        else:
            current_url = a_row[URL_COLUMN]
            if current_url == '':
                continue
            if a_idx % 10 == 0:
                url_words = current_url.split('/')
                url_end = url_words[-1].split('&')[0]
                true_url = "https://acs.leagueoflegends.com/v1/stats/game/" + url_words[-2] + "/" + url_end

                r = requests.get(true_url)
                match_detail = json.loads(r.text)
            damage_taken = match_detail['participants'][a_idx % 10]['stats']['totalDamageTaken']
            #print str(a_idx) + " - other damage taken:" + str(damage_taken)
        if a_idx % 100 == 0:
            print "进度%d/%d" % (a_idx, total_num)
        if current_league == "LPL":
            a_row[DAMAGE_TO_CHAMPIONS_COLUMN] = lpl_damage_to_champion
            a_row[TOTAL_GOLD_COLUMN] = gold_earned

        a_row.append(damage_taken)
        
        csvwriter.writerow(a_row)


# In[86]:

def AddToDict(a_dict, a_key, a_value):
    if a_key in a_dict.keys():
        a_dict[a_key] += a_value
    else:
        a_dict[a_key] = a_value


# In[97]:

print "开始统计选手数据"
player_dpm_sum_dict = {}
player_gpm_sum_dict = {}
player_num_dict = {}
player_ka_damage_sum_dict = {}
player_d_damage_taken_sum_dict = {}
player_damage_to_gold_sum_dict = {}
player_damage_to_gold_share_sum_dict = {}
player_damage_share_sum_dict = {}
player_damage_taken_share_sum_dict = {}
player_kill_sum_dict = {}
player_death_sum_dict = {}
player_assist_sum_dict = {}
player_win_sum_dict = {}
player_champs_dict = {}
player_league_dict = {}
player_position_dict = {}

temp_total_damage_taken_team = 0
temp_total_damage_team = 0
temp_total_gold_team = 0

for a_idx, a_row in enumerate(all_region_table):
    a_player = a_row[PLAYER_COLUMN]
    a_position = a_row[POSITION_COLUMN]
    a_team = a_row[TEAM_COLUMN]
    a_champ = a_row[CHAMPION_COLUMN]
    a_total_damage = float(a_row[DAMAGE_TO_CHAMPIONS_COLUMN])
    a_game_length = float(a_row[GAMELENGTH_COLUMN])
    a_dpm = a_total_damage / a_game_length
    a_total_gold = float(a_row[TOTAL_GOLD_COLUMN])
    a_gpm = (a_total_gold - 500) / a_game_length #500 is the starting gold
    
    AddToDict(player_dpm_sum_dict, a_player, a_dpm)
    AddToDict(player_gpm_sum_dict, a_player, a_gpm)
    AddToDict(player_num_dict, a_player, 1)
    
    
    a_kill = int(a_row[KILL_COLUMN])
    a_death = int(a_row[DEATH_COLUMN])
    a_assist = int(a_row[ASSIST_COLUMN])

    a_average_ka_damage = 0
    if a_kill + a_assist == 0:
        a_average_ka_damage = float(a_total_damage)
    else:
        a_average_ka_damage = float(a_total_damage) / float(int(a_kill) + int(a_assist))
    AddToDict(player_ka_damage_sum_dict, a_player, a_average_ka_damage)
    
    a_total_damage_taken = float(a_row[CUSTOM_DAMAGE_TAKEN_COLUMN])
    a_average_d_damage_taken = 0
    if a_death == 0:
        a_average_d_damage_taken = float(a_total_damage_taken)
    else:
        a_average_d_damage_taken = float(a_total_damage_taken) / float(int(a_death))
    AddToDict(player_d_damage_taken_sum_dict, a_player, a_average_d_damage_taken)
    
    if a_idx % 5 == 0:
        temp_total_damage_team = 0
        temp_total_gold_team = 0
        for team_idx in range(5):
            temp_total_damage_team += float(all_region_table[a_idx + team_idx][DAMAGE_TO_CHAMPIONS_COLUMN])
            temp_total_gold_team += float(all_region_table[a_idx + team_idx][TOTAL_GOLD_COLUMN])
    
    if a_total_gold == 0:
        print "发现一处数据错误，详细信息——战队:%s 选手:%s 英雄:%s 位置:%s，忽略该条数据" % (a_team, a_player, a_champ, a_position)
        continue
    a_gold_share = a_total_gold / temp_total_gold_team
    a_damage_share = float(a_total_damage) / temp_total_damage_team
    a_damage_to_gold = a_total_damage / float(a_total_gold)
    a_damage_to_gold_share = a_damage_share / float(a_gold_share)
    AddToDict(player_damage_to_gold_sum_dict, a_player, a_damage_to_gold)
    AddToDict(player_damage_to_gold_share_sum_dict, a_player, a_damage_to_gold_share)
    AddToDict(player_damage_share_sum_dict, a_player, a_damage_share)
    
    if a_idx % 5 == 0:
        temp_total_damage_taken_team = 0
        for team_idx in range(5):
            temp_total_damage_taken_team += float(all_region_table[a_idx + team_idx][CUSTOM_DAMAGE_TAKEN_COLUMN])
    a_damage_taken_share = a_total_damage_taken / temp_total_damage_taken_team
    AddToDict(player_damage_taken_share_sum_dict, a_player, a_damage_taken_share)
    
    AddToDict(player_kill_sum_dict, a_player, a_kill)
    AddToDict(player_death_sum_dict, a_player, a_death)
    AddToDict(player_assist_sum_dict, a_player, a_assist)
    
    a_win = float(a_row[RESULT_COLUMN])
    AddToDict(player_win_sum_dict, a_player, a_win)
    
    a_champ = a_row[CHAMPION_COLUMN]
    if a_player in player_champs_dict.keys():
        AddToDict(player_champs_dict[a_player], a_champ, 1)
    else:
        player_champs_dict[a_player] = {}
        AddToDict(player_champs_dict[a_player], a_champ, 1)
        
    player_league_dict[a_player] = a_row[LEAGUE_COLUMN]
    player_position_dict[a_player] = a_row[POSITION_COLUMN]
    
    #print a_player, a_team, a_dpm, a_gpm, a_average_ka_damage, a_average_d_damage_taken
    if a_idx % 100 == 0:
        print "进度%d/%d" % (a_idx, total_num)


# In[99]:

print "将选手数据写入文件" + (RESULT_FOLDER + output_player_file_name) % latest_week
output_player_def_row = [u"选手",u"每分钟对英雄伤害", u"每分钟金钱", u"平均每击败需要伤害", u"平均每死亡承受伤害",                         u"伤害金钱比", u"伤害金钱比（占团队比例）", u"伤害占团队百分比", u"承伤占团队百分比",                         u"击杀", u"死亡", u"助攻", u"胜率", u"选取英雄频次", u"赛区", u"位置", u"出场次数"]
with open((RESULT_FOLDER + output_player_file_name) % latest_week, 'wb') as csvfile:
    csvfile.write(u'\ufeff'.encode('utf8'))
    csvwriter = csv.writer(csvfile, delimiter=',',
                            quotechar='|', quoting=csv.QUOTE_MINIMAL)
    csvwriter.writerow([item.encode('utf8') for item in output_player_def_row])
    
    for a_player in player_num_dict:
        csvwriter.writerow([a_player, 
                            player_dpm_sum_dict[a_player]/float(player_num_dict[a_player]),
                            player_gpm_sum_dict[a_player]/float(player_num_dict[a_player]),
                            player_ka_damage_sum_dict[a_player]/float(player_num_dict[a_player]),
                            player_d_damage_taken_sum_dict[a_player]/float(player_num_dict[a_player]),
                            player_damage_to_gold_sum_dict[a_player]/float(player_num_dict[a_player]),
                            player_damage_to_gold_share_sum_dict[a_player]/float(player_num_dict[a_player]),
                            player_damage_share_sum_dict[a_player]/float(player_num_dict[a_player]),
                            player_damage_taken_share_sum_dict[a_player]/float(player_num_dict[a_player]),
                            player_kill_sum_dict[a_player]/float(player_num_dict[a_player]),
                            player_death_sum_dict[a_player]/float(player_num_dict[a_player]),
                            player_assist_sum_dict[a_player]/float(player_num_dict[a_player]),
                            player_win_sum_dict[a_player]/float(player_num_dict[a_player]),
                            ' '.join('{}-{}'.format(key, val) for key, val in sorted(player_champs_dict[a_player].items())),
                            player_league_dict[a_player],
                            player_position_dict[a_player],
                            player_num_dict[a_player]
                           ])


# In[100]:

print "开始统计英雄数据"
champ_dpm_sum_dict = {}
champ_gpm_sum_dict = {}
champ_num_dict = {}
champ_ka_damage_sum_dict = {}
champ_d_damage_taken_sum_dict = {}
champ_damage_to_gold_sum_dict = {}
champ_damage_to_gold_share_sum_dict = {}
champ_damage_share_sum_dict = {}
champ_damage_taken_share_sum_dict = {}
champ_kill_sum_dict = {}
champ_death_sum_dict = {}
champ_assist_sum_dict = {}
champ_win_sum_dict = {}
champ_champs_dict = {}
champ_league_dict = {}

champ_ban_sum_dict = {}
champ_ban1_sum_dict = {}
champ_ban2_sum_dict = {}
champ_pick_sum_dict = {}

temp_total_damage_taken_team = 0
temp_total_damage_team = 0
temp_total_gold_team = 0

for a_idx, a_row in enumerate(all_region_table):
    a_player = a_row[PLAYER_COLUMN]
    a_position = a_row[POSITION_COLUMN]
    a_team = a_row[TEAM_COLUMN]
    a_champ = a_row[CHAMPION_COLUMN]
    
    a_champ_position_key = a_champ + '_' + a_position
    
    a_total_damage = float(a_row[DAMAGE_TO_CHAMPIONS_COLUMN])
    a_game_length = float(a_row[GAMELENGTH_COLUMN])
    a_dpm = a_total_damage / a_game_length
    a_total_gold = float(a_row[TOTAL_GOLD_COLUMN])
    a_gpm = (a_total_gold - 500) / a_game_length #500 is the starting gold

    if a_total_gold == 0:
        print "发现一处数据错误，详细信息——战队:%s 选手:%s 英雄:%s 位置:%s，忽略该条数据" % (a_team, a_player, a_champ, a_position)
        continue
    
    AddToDict(champ_dpm_sum_dict, a_champ_position_key, a_dpm)
    AddToDict(champ_gpm_sum_dict, a_champ_position_key, a_gpm)
    AddToDict(champ_num_dict, a_champ_position_key, 1)
    
    
    a_kill = int(a_row[KILL_COLUMN])
    a_death = int(a_row[DEATH_COLUMN])
    a_assist = int(a_row[ASSIST_COLUMN])

    a_average_ka_damage = 0
    if a_kill + a_assist == 0:
        a_average_ka_damage = float(a_total_damage)
    else:
        a_average_ka_damage = float(a_total_damage) / float(int(a_kill) + int(a_assist))
    AddToDict(champ_ka_damage_sum_dict, a_champ_position_key, a_average_ka_damage)
    
    a_total_damage_taken = float(a_row[CUSTOM_DAMAGE_TAKEN_COLUMN])
    a_average_d_damage_taken = 0
    if a_death == 0:
        a_average_d_damage_taken = float(a_total_damage_taken)
    else:
        a_average_d_damage_taken = float(a_total_damage_taken) / float(int(a_death))
    AddToDict(champ_d_damage_taken_sum_dict, a_champ_position_key, a_average_d_damage_taken)
    
    if a_idx % 5 == 0:
        temp_total_damage_team = 0
        temp_total_gold_team = 0
        for team_idx in range(5):
            temp_total_damage_team += float(all_region_table[a_idx + team_idx][DAMAGE_TO_CHAMPIONS_COLUMN])
            temp_total_gold_team += float(all_region_table[a_idx + team_idx][TOTAL_GOLD_COLUMN])

    a_gold_share = a_total_gold / temp_total_gold_team
    a_damage_share = float(a_total_damage) / temp_total_damage_team
    a_damage_to_gold = a_total_damage / float(a_total_gold)
    a_damage_to_gold_share = a_damage_share / float(a_gold_share)
    AddToDict(champ_damage_to_gold_sum_dict, a_champ_position_key, a_damage_to_gold)
    AddToDict(champ_damage_to_gold_share_sum_dict, a_champ_position_key, a_damage_to_gold_share)
    AddToDict(champ_damage_share_sum_dict, a_champ_position_key, a_damage_share)
    
    if a_idx % 5 == 0:
        temp_total_damage_taken_team = 0
        for team_idx in range(5):
            temp_total_damage_taken_team += float(all_region_table[a_idx + team_idx][CUSTOM_DAMAGE_TAKEN_COLUMN])
    a_damage_taken_share = a_total_damage_taken / temp_total_damage_taken_team
    AddToDict(champ_damage_taken_share_sum_dict, a_champ_position_key, a_damage_taken_share)
    
    AddToDict(champ_kill_sum_dict, a_champ_position_key, a_kill)
    AddToDict(champ_death_sum_dict, a_champ_position_key, a_death)
    AddToDict(champ_assist_sum_dict, a_champ_position_key, a_assist)
    
    a_win = float(a_row[RESULT_COLUMN])
    AddToDict(champ_win_sum_dict, a_champ_position_key, a_win)
    
    if a_idx % 5 == 0:
        AddToDict(champ_ban_sum_dict, a_row[BAN1_COLUMN], 1)
        AddToDict(champ_ban_sum_dict, a_row[BAN2_COLUMN], 1)
        AddToDict(champ_ban_sum_dict, a_row[BAN3_COLUMN], 1)
        AddToDict(champ_ban_sum_dict, a_row[BAN4_COLUMN], 1)
        AddToDict(champ_ban_sum_dict, a_row[BAN5_COLUMN], 1)
        
        AddToDict(champ_ban1_sum_dict, a_row[BAN1_COLUMN], 1)
        AddToDict(champ_ban1_sum_dict, a_row[BAN2_COLUMN], 1)
        AddToDict(champ_ban1_sum_dict, a_row[BAN3_COLUMN], 1)
        
        AddToDict(champ_ban2_sum_dict, a_row[BAN4_COLUMN], 1)
        AddToDict(champ_ban2_sum_dict, a_row[BAN5_COLUMN], 1)
    if a_idx % 100 == 0:
        print "进度%d/%d" % (a_idx, total_num)


# In[103]:

output_champion_def_row = [u"英雄", u"位置", u"每分钟对英雄伤害", u"每分钟金钱", u"平均每击败需要伤害", u"平均每死亡承受伤害",                         u"伤害金钱比", u"伤害金钱比（占团队比例）", u"伤害占团队百分比", u"承伤占团队百分比",                         u"击杀", u"死亡", u"助攻", u"胜率", u"BAN", u"第一轮ban", u"第二轮ban", u"PICK", u"BAN+PICK"]
print "将英雄数据写入文件" + (RESULT_FOLDER + output_champion_file_name) % latest_week
match_num = len(all_region_table) / 10
with open((RESULT_FOLDER + output_champion_file_name) % latest_week, 'wb') as csvfile:
    csvfile.write(u'\ufeff'.encode('utf8'))
    csvwriter = csv.writer(csvfile, delimiter=',',
                            quotechar='|', quoting=csv.QUOTE_MINIMAL)
    csvwriter.writerow([item.encode('utf8') for item in output_champion_def_row])
    
    for a_champ_position in champ_num_dict:
        a_champ = a_champ_position.split('_')[0]
        a_position = a_champ_position.split('_')[1]
        
        ban_rate = 0
        ban1_rate = 0
        ban2_rate = 0
        pick_rate = 0
        if a_champ in champ_ban_sum_dict.keys():
            ban_rate = champ_ban_sum_dict[a_champ]#/float(match_num)
        if a_champ_position in champ_num_dict.keys():
            pick_rate = champ_num_dict[a_champ_position]#/float(match_num)
        if a_champ in champ_ban1_sum_dict.keys():
            ban1_rate = champ_ban1_sum_dict[a_champ]
        if a_champ in champ_ban2_sum_dict.keys():
            ban2_rate = champ_ban2_sum_dict[a_champ]
        csvwriter.writerow([a_champ, 
                            a_position,
                            champ_dpm_sum_dict[a_champ_position]/float(champ_num_dict[a_champ_position]),
                            champ_gpm_sum_dict[a_champ_position]/float(champ_num_dict[a_champ_position]),
                            champ_ka_damage_sum_dict[a_champ_position]/float(champ_num_dict[a_champ_position]),
                            champ_d_damage_taken_sum_dict[a_champ_position]/float(champ_num_dict[a_champ_position]),
                            champ_damage_to_gold_sum_dict[a_champ_position]/float(champ_num_dict[a_champ_position]),
                            champ_damage_to_gold_share_sum_dict[a_champ_position]/float(champ_num_dict[a_champ_position]),
                            champ_damage_share_sum_dict[a_champ_position]/float(champ_num_dict[a_champ_position]),
                            champ_damage_taken_share_sum_dict[a_champ_position]/float(champ_num_dict[a_champ_position]),
                            champ_kill_sum_dict[a_champ_position]/float(champ_num_dict[a_champ_position]),
                            champ_death_sum_dict[a_champ_position]/float(champ_num_dict[a_champ_position]),
                            champ_assist_sum_dict[a_champ_position]/float(champ_num_dict[a_champ_position]),
                            champ_win_sum_dict[a_champ_position]/float(champ_num_dict[a_champ_position]),
                            ban_rate,
                            ban1_rate,
                            ban2_rate,
                            pick_rate,
                            ban_rate + pick_rate,
                           ])
        
print "脚本完成，选手及英雄数据记录在目录%s中的%s和%s文件" % (RESULT_FOLDER, output_player_file_name % latest_week, output_champion_file_name % latest_week)


# In[ ]:



